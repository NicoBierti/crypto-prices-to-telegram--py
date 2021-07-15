# Crypto Prices to Telegram Bot
# See README file before running the code.

from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import datetime
import time
import telegram
import cryptocompare

#To calculate the time execution of the script
begin_time = time.time()

# Set working directory and workbook

os.chdir(os.getcwd())
file_name = 'Crypto_Prices.xlsx'
workbook = load_workbook(file_name)
worksheet = workbook.active

#Formating
worksheet['A1'] = 'Date'
worksheet['B1'] = 'Time'  
worksheet['C1'] = 'BTC (USD)'
worksheet['D1'] = 'ETH (USD)'
worksheet['E1'] = 'DOGE (USD)'
worksheet['F1'] = 'ADA (USD)'
worksheet['G1'] = 'BNB (USD)'
worksheet['H1'] = 'CAKE (USD)'

bold = Font( bold = True)
for cell in worksheet["1:1"]:
    cell.font = bold     

# Getting Time and Date
now = datetime.datetime.now()
date = (now.strftime("%d/%m/%y"))
current_time = (now.strftime("%X"))

# Function to write cells
def write_data(columns, data, max_row_parameter):
    for rows in range (1 , worksheet.max_row + max_row_parameter):  
        if worksheet.cell(row = rows, column = columns).value != None:
            continue
        else:
            worksheet.cell(row = rows, column = columns).value = data

# Writes "Current Date" in column 1
write_data(1,date,2)
# Writes "Current Time" in column 2
write_data(2,current_time,1)

# Get Cryptocurrencies prices in USD using Cryptocompare.com API.
cryptocompare.cryptocompare._set_api_key_parameter('##################################')

btc_price_usd = cryptocompare.get_price(['BTC'],['USD'])['BTC']['USD']
eth_price_usd =cryptocompare.get_price(['ETH'],['USD'])['ETH']['USD']
doge_price_usd = cryptocompare.get_price(['DOGE'],['USD'])['DOGE']['USD']
ada_price_usd =cryptocompare.get_price(['ADA'],['USD'])['ADA']['USD']
bnb_price_usd =cryptocompare.get_price(['BNB'],['USD'])['BNB']['USD']
cake_price = cryptocompare.get_price(['CAKE'],['USD'])
cake_price_usd = cryptocompare.get_price(['CAKE'],['USD'])['CAKE']['USD']

# Writes Crypto Prices Excel
write_data(3,btc_price_usd,1)
write_data(4,eth_price_usd,1)
write_data(5,doge_price_usd,1)
write_data(6,ada_price_usd,1)
write_data(7,bnb_price_usd,1)
write_data(8,cake_price_usd,1)

# Function to send message to Telegram Bot
my_token = 'XXXXXXXXXXXX'   # Paste token
chat_id = int('XXXXXXXXXXXXX') # Paste chat_id

def send(message, chat_id, token = my_token):
	
	bot = telegram.Bot(token)
	bot.sendMessage(chat_id = chat_id, text = message)
 
# Message content
send('Today ' + str(date) + ' at ' + str(current_time) + ' Crypto prices are: ' , chat_id)
send('BTC = ' + str(btc_price_usd) + ' USD' , chat_id)
send('ETH = ' + str(eth_price_usd) + ' USD' , chat_id)
send('DOGE = ' + str(doge_price_usd) + ' USD' , chat_id)
send('ADA = ' + str(ada_price_usd) + ' USD' , chat_id)
send('BNB = ' + str(bnb_price_usd) + ' USD' , chat_id)
send('CAKE = ' + str(cake_price_usd) + ' USD' , chat_id)

# Saves the file and calculates script execution time
workbook.save(file_name)       
end_time = time.time()
print('Execution time is', (end_time - begin_time))